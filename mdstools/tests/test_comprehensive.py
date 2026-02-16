#!/usr/bin/env python
"""
Comprehensive test suite for metadata flattening and schema enrichment.

This demonstrates the complete workflow from YAML to enriched Excel/CSV files.
"""

import json
import os
from pathlib import Path

from mdstools.metadata.enriched_metadata import EnrichedFlattenedMetadata
from mdstools.metadata.flattened_metadata import FlattenedMetadata
from mdstools.metadata.metadata import Metadata


def test_basic_flattening():
    """Test basic YAML flattening without schema enrichment."""
    print("\n" + "=" * 80)
    print("TEST 1: Basic Flattening")
    print("=" * 80)

    # Load test data from YAML
    metadata = Metadata.from_yaml("tests/simple_test.yaml")

    # Flatten to tabular format
    flattened = metadata.flatten()

    # Check flattened output
    df = flattened.to_pandas()
    print(f"✓ Flattened to {len(df)} rows")
    print(f"✓ Columns: {list(df.columns)}")

    # Export to CSV
    output_dir = Path("tests/generated/test_basic_flattening")
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "basic_flat.csv"
    flattened.to_csv(str(output))
    print(f"✓ Exported to {output}")

    assert output.exists()
    assert len(df) > 0
    assert list(df.columns) == ["Number", "Key", "Value"]


def test_schema_enrichment():
    """Test schema-based enrichment with descriptions and examples."""
    print("\n" + "=" * 80)
    print("TEST 2: Schema Enrichment")
    print("=" * 80)

    # Load test data from YAML
    metadata = Metadata.from_yaml("tests/simple_test.yaml")

    # Flatten to tabular format
    flattened = metadata.flatten()

    # Enrich with schema information
    enriched = EnrichedFlattenedMetadata(flattened.rows, schema_dir="schemas")

    # Check enriched output
    df = enriched.to_pandas()
    print(f"✓ Enriched to {len(df)} rows")
    print(f"✓ Columns: {list(df.columns)}")

    # Check that some descriptions were added
    has_desc = df["Description"].notna() & (df["Description"] != "")
    desc_count = has_desc.sum()
    print(f"✓ Found {desc_count} fields with descriptions")

    # Export enriched CSV
    output_dir = Path("tests/generated/test_schema_enrichment")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_csv = output_dir / "enriched.csv"
    enriched.to_csv(str(output_csv))
    print(f"✓ Exported enriched CSV to {output_csv}")

    # Export enriched Excel
    output_xlsx = output_dir / "enriched.xlsx"
    enriched.to_excel(str(output_xlsx))
    print(f"✓ Exported enriched Excel to {output_xlsx}")

    assert output_csv.exists()
    assert output_xlsx.exists()
    assert list(df.columns) == ["Number", "Key", "Value", "Example", "Description"]
    assert desc_count > 0


def test_multi_sheet_export():
    """Test export to multi-sheet Excel file."""
    print("\n" + "=" * 80)
    print("TEST 3: Multi-Sheet Excel Export")
    print("=" * 80)

    # Load test data from YAML
    metadata = Metadata.from_yaml("tests/simple_test.yaml")

    # Flatten and enrich
    flattened = metadata.flatten()
    enriched = EnrichedFlattenedMetadata(flattened.rows, schema_dir="schemas")

    # Export to multi-sheet Excel
    output_dir = Path("tests/generated/test_multi_sheet_export")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_multi = output_dir / "enriched_multi_sheet.xlsx"
    enriched.to_excel(str(output_multi), separate_sheets=True)
    print(f"✓ Exported multi-sheet Excel to {output_multi}")

    # Also test with FlattenedMetadata
    output_flat_multi = output_dir / "flattened_multi_sheet.xlsx"
    flattened.to_excel(str(output_flat_multi), separate_sheets=True)
    print(f"✓ Exported multi-sheet Excel (flattened) to {output_flat_multi}")

    # Verify files exist
    assert output_multi.exists()
    assert output_flat_multi.exists()

    # Load and verify the multi-sheet file has multiple sheets
    import openpyxl

    wb = openpyxl.load_workbook(output_multi)
    sheet_count = len(wb.sheetnames)
    print(f"✓ Multi-sheet Excel has {sheet_count} sheet(s): {wb.sheetnames}")
    wb.close()

    assert sheet_count > 0, "Excel file should have at least one sheet"

    # Test roundtrip: save multi-sheet → load → verify data matches
    print("\n" + "Testing multi-sheet roundtrip...")
    loaded_enriched = EnrichedFlattenedMetadata.from_excel(
        str(output_multi), schema_dir="schemas"
    )
    loaded_flattened = FlattenedMetadata.from_excel(str(output_flat_multi))

    # Verify loaded data matches original (compare base rows, not enrichment)
    assert len(loaded_enriched.base_rows) == len(enriched.base_rows)
    assert len(loaded_flattened.rows) == len(flattened.rows)
    print("✓ Multi-sheet roundtrip successful - data preserved")

    # Verify unflattening still works correctly
    original_data = metadata.data
    loaded_data = loaded_flattened.unflatten().data
    assert loaded_data == original_data
    print("✓ Multi-sheet data correctly unflattens to original structure")


def test_specific_field_enrichment():
    """Test that specific fields get correct enrichment."""
    print("\n" + "=" * 80)
    print("TEST 4: Specific Field Enrichment")
    print("=" * 80)

    # Load test data from YAML
    metadata = Metadata.from_yaml("tests/simple_test.yaml")

    # Flatten and enrich
    flattened = metadata.flatten()
    enriched = EnrichedFlattenedMetadata(flattened.rows, schema_dir="schemas")
    df = enriched.to_pandas()

    # Check specific fields
    test_cases = [
        ("role", "curation.process.role"),
        ("name", "curation.process.name"),
        ("type", "system.type"),
    ]

    for key, expected_path in test_cases:
        rows = df[df["Key"] == key]
        if not rows.empty:
            first_match = rows.iloc[0]
            if first_match["Description"]:
                print(
                    f"✓ '{key}' has description: {first_match['Description'][:50]}..."
                )
            if first_match["Example"]:
                print(f"  Example: {first_match['Example']}")


def test_markdown_export():
    """Test markdown export functionality."""
    print("\n" + "=" * 80)
    print("TEST 5: Markdown Export")
    print("=" * 80)

    # Load test data from YAML
    metadata = Metadata.from_yaml("tests/simple_test.yaml")

    # Flatten and enrich
    flattened = metadata.flatten()
    enriched = EnrichedFlattenedMetadata(flattened.rows, schema_dir="schemas")

    # Export to markdown
    markdown = enriched.to_markdown()

    # Save to file
    output_dir = Path("tests/generated/test_markdown_export")
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "enriched.md"
    with open(output, "w", encoding="utf-8") as f:
        f.write(markdown)

    print(f"✓ Exported markdown to {output}")
    print(f"✓ Markdown length: {len(markdown)} characters")

    assert len(markdown) > 0
    assert "|" in markdown  # Should be a table
    assert "Description" in markdown


def test_unflatten_validation_and_cli():
    """Test schema validation on unflatten and CLI unflatten conversion."""
    print("\n" + "=" * 80)
    print("TEST 6: Unflatten Validation + CLI")
    print("=" * 80)

    # Use the comprehensive example metadata
    output_dir = Path("tests/generated/test_unflatten_validation_and_cli")
    output_dir.mkdir(parents=True, exist_ok=True)

    metadata = Metadata.from_yaml("tests/simple_test.yaml")
    flattened = metadata.flatten()

    output_xlsx = output_dir / "unflatten_validation.xlsx"
    flattened.to_excel(str(output_xlsx))

    schema_file = output_dir / "unflatten_schema.json"
    schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "type": "object",
        "required": ["curation", "system"],
        "additionalProperties": True,
        "properties": {
            "curation": {"type": "object", "additionalProperties": True},
            "system": {"type": "object", "additionalProperties": True},
        },
    }
    with open(schema_file, "w", encoding="utf-8") as f:
        json.dump(schema, f)

    loaded = FlattenedMetadata.from_excel(str(output_xlsx))
    validated = loaded.unflatten(schema_path=str(schema_file))
    assert "curation" in validated.data
    assert "system" in validated.data

    # CLI unflatten with validation
    from mdstools import cli as md_cli

    cli_output_dir = output_dir / "cli_unflatten"
    result = md_cli.main(
        [
            "unflatten",
            str(output_xlsx),
            "--out-dir",
            str(cli_output_dir),
            "--schema-file",
            str(schema_file),
        ]
    )

    output_yaml = cli_output_dir / f"{output_xlsx.stem}.yaml"
    assert result == 0
    assert output_yaml.exists()


def run_all_tests():
    """Run all tests."""
    print("\n" + "=" * 80)
    print("RUNNING COMPREHENSIVE TEST SUITE")
    print("=" * 80)

    # Ensure output directory exists
    os.makedirs("tests/generated", exist_ok=True)

    try:
        test_basic_flattening()
        test_schema_enrichment()
        test_multi_sheet_export()
        test_specific_field_enrichment()
        test_markdown_export()
        test_unflatten_validation_and_cli()

        print("\n" + "=" * 80)
        print("✓ ALL TESTS PASSED")
        print("=" * 80)
        print("\nGenerated test files in tests/generated/:")
        for f in Path("tests/generated").iterdir():
            if f.is_file():
                print(f"  - {f.name} ({f.stat().st_size} bytes)")

    except Exception as e:
        print(f"\n✗ TEST FAILED: {e}")
        raise


if __name__ == "__main__":
    run_all_tests()
